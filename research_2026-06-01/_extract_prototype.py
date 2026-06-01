"""Извлечь структуру prototype_table.xlsx в текстовый файл для анализа."""
import openpyxl
from pathlib import Path

src = Path("e:/Arhiv/projects/MyCash/prototype_table.xlsx")
dst = Path("e:/Arhiv/projects/MyCash/research_2026-06-01/_prototype_dump.md")

wb = openpyxl.load_workbook(str(src), data_only=False)

lines = [f"# prototype_table.xlsx — структура", ""]
lines.append(f"Листы ({len(wb.sheetnames)}): {wb.sheetnames}")
lines.append("")

for name in wb.sheetnames:
    ws = wb[name]
    lines.append(f"## Лист «{name}»  ({ws.max_row} строк, {ws.max_column} колонок)")
    lines.append("")

    # Выгружаем первые 12 строк (или меньше) — для понимания структуры
    rows_to_dump = min(ws.max_row, 12)
    for r in range(1, rows_to_dump + 1):
        row_cells = []
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is None:
                v = ""
            else:
                v = str(v)
                if len(v) > 40:
                    v = v[:37] + "..."
            row_cells.append(v)
        # Убираем хвостовые пустые
        while row_cells and row_cells[-1] == "":
            row_cells.pop()
        if row_cells:
            lines.append(f"  R{r}: " + " | ".join(row_cells))
        else:
            lines.append(f"  R{r}: (пусто)")
    if ws.max_row > rows_to_dump:
        lines.append(f"  ... ещё {ws.max_row - rows_to_dump} строк")
    lines.append("")

dst.write_text("\n".join(lines), encoding="utf-8")
print(f"Saved: {dst}")
print(f"Size: {dst.stat().st_size} bytes")
