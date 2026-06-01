"""Извлечь полный список статей ДДС из листа «ДДС статьи».

Также вытащить направления из листа «Справочники».
"""
import openpyxl
from pathlib import Path
import json

src = Path("e:/Arhiv/projects/MyCash/prototype_table.xlsx")
dst_md = Path("e:/Arhiv/projects/MyCash/research_2026-06-01/_dds_articles_dump.md")
dst_json = Path("e:/Arhiv/projects/MyCash/research_2026-06-01/_dds_articles_dump.json")

wb = openpyxl.load_workbook(str(src), data_only=True)

# ===== Лист «ДДС статьи» =====
ws = wb["ДДС статьи"]
print(f"Лист «ДДС статьи»: {ws.max_row} строк")

# Структура колонок: Статья ДДС | Группа | Вид деятельности | Субстатья ДДС
articles = []
for r in range(2, ws.max_row + 1):
    name = ws.cell(row=r, column=1).value
    group = ws.cell(row=r, column=2).value
    activity = ws.cell(row=r, column=3).value
    subarticle = ws.cell(row=r, column=4).value
    if name is None and group is None and activity is None:
        continue
    articles.append({
        "name": (name or "").strip() if name else "",
        "group": (group or "").strip() if group else "",
        "activity": (activity or "").strip() if activity else "",
        "subarticle": (subarticle or "").strip() if subarticle else "",
    })

# ===== Лист «Справочники» — направления =====
ws2 = wb["Справочники"]
directions = []
for r in range(2, min(ws2.max_row + 1, 100)):
    direction = ws2.cell(row=r, column=1).value
    if direction:
        directions.append(str(direction).strip())

# ===== Лист «ДДС настройки (для ввода сальдо» — стартовое сальдо/кошельки =====
ws3 = None
for name in wb.sheetnames:
    if "настройки" in name.lower():
        ws3 = wb[name]
        break

wallet_seeds = []
if ws3:
    # Структура: Месяц начала | <число>; затем Кошелек | Сумма; затем строки кошельков
    for r in range(2, ws3.max_row + 1):
        kw = ws3.cell(row=r, column=1).value
        amount = ws3.cell(row=r, column=2).value
        if kw and isinstance(kw, str) and kw not in ("Кошелек",):
            wallet_seeds.append({"name": kw.strip(), "balance": amount or 0})

# ===== Сохранить =====
data = {
    "articles_count": len(articles),
    "articles": articles,
    "directions_seed": directions,
    "wallets_seed": wallet_seeds,
}
dst_json.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

lines = ["# Реальные данные prototype_table.xlsx для seed", ""]
lines.append(f"## Кошельки ({len(wallet_seeds)})")
for w in wallet_seeds:
    lines.append(f"- {w['name']} (стартовое сальдо: {w['balance']})")
lines.append("")
lines.append(f"## Направления ({len(directions)})")
for d in directions:
    lines.append(f"- {d}")
lines.append("")
lines.append(f"## Статьи ДДС ({len(articles)})")
lines.append("")
lines.append("| # | Статья | Группа | Вид деятельности | Субстатья |")
lines.append("|---|--------|--------|------------------|-----------|")
for i, a in enumerate(articles, start=1):
    lines.append(f"| {i} | {a['name']} | {a['group']} | {a['activity']} | {a['subarticle']} |")

dst_md.write_text("\n".join(lines), encoding="utf-8")
print(f"Saved: {dst_md}")
print(f"Saved: {dst_json}")
print(f"Articles: {len(articles)}, Directions: {len(directions)}, Wallets: {len(wallet_seeds)}")
